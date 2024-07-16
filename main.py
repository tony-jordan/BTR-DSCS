# Dependencies
import tkinter as tk
from tkinter import Toplevel
from PIL import Image, ImageTk
from location_button import location_btn
import win32com.client as win32
from openpyxl import load_workbook
import os

# Constants
# IT hub receiving areas
spare_locations = [
    "Spare.IT Cage.Loaner Rack.0.0.0.0.0.0.0",
    "Spare.IT.Available.Deployment Rack.0.0.0.0.0.0",         
    "Spare.IT.Available.EC Rack.0.0.0.0.0.0",                 
    "Spare.IT.Available.Rack 2.0.0.0.0.0.0",
    "Spare.IT.Available.Rack 3.0.0.0.0.0.0",
    "Spare.IT.Available.Rack 4.0.0.0.0.0.0",
    "Spare.IT.Available.Rack 5.0.0.0.0.0.0",
    "Spare.IT.Available.Rack 6.0.0.0.0.0.0",                  
    "Spare.IT.Available.Rack 7.0.0.0.0.0.0",                 
    "Spare.IT.Available.Radio Rack.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf A.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf B.0.0.0.0.0.0",                 
    "Spare.IT.Available.Shelf C.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf D.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf E.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf F.0.0.0.0.0.0",                 
    "Spare.IT.Available.Shelf G.0.0.0.0.0.0",                
    "Spare.IT.Available.Shelf H.0.0.0.0.0.0",                 
    "Spare.IT.Available.Shelf I.0.0.0.0.0.0",                 
    "Spare.IT.Available.Shelf J.0.0.0.0.0.0",                
    "Spare.IT.Available.Shelf K.0.0.0.0.0.0",                
    "Spare.IT.Available.Shelf L.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf M.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf N.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf O.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf P.0.0.0.0.0.0",
    "Spare.IT.Available.Shelf Q.0.0.0.0.0.0",               
    "Spare.IT.Available.Shelf R.0.0.0.0.0.0",              
]

# Warehouse department locations
in_use_locations = [
    #!!INTERNAL DATA REDACTED!!
]
alias_names = []

active_location_id = [-1]
active_mode = [0]
active_windows = [0]

# Function to update/create excel spreadsheet with inventory move data
def update_excel(id, assets):
    workbook = load_workbook(filename="BOT_Format_11.22.23.xlsx")
    sheet = workbook.active
    for x, asset in enumerate(assets):
	# format ind[ex] for writing to excel spreadsheet
        ind = str(x + 2)
        sheet["A" + ind] = asset
        location_array = []
        pointr = ''
	// format locations in excel spreadsheet
        if active_mode[0] == 0:
            for chr in in_use_locations[id]:
                if chr == '.':
                    location_array.append(pointr)
                    pointr = ''
                else:
                    pointr += chr
        else:
            for chr in spare_locations[id]:
                if chr == '.':
                    location_array.append(pointr)
                    pointr = ''
                else:
                    pointr += chr
	// continue formatting for excel spreadsheet
        sheet["D" + ind] = location_array[0]
        sheet["E" + ind] = "No"
        sheet["F" + ind] = location_array[1]
        sheet["G" + ind] = location_array[2]
        sheet["H" + ind] = location_array[3]
        sheet["I" + ind] = location_array[4]
        sheet["J" + ind] = location_array[5]
        sheet["K" + ind] = location_array[6]
        sheet["L" + ind] = ""
        sheet["M" + ind] = "user login" #!!INTERNAL DATA REDACTED

    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    workbook.save(filename=ROOT_DIR + "\sendMe.xlsx")
    send_mail()                                            

# function to update location to assign equipment to specific users
def update_excel_user(asset, user):
    workbook = load_workbook(filename="BOT_Format_11.22.23.xlsx")
    sheet = workbook.active
    sheet["A2"] = asset
    sheet["D2"] = "In-Use"
    sheet["E2"] = "Yes"
    sheet["F2"] = user
    sheet["G2"] = "0"
    sheet["H2"] = "0"
    sheet["I2"] = "0"
    sheet["J2"] = "0"
    sheet["K2"] = "0"
    sheet["L2"] = "09/09/9999"
    sheet["M2"] = user

    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    workbook.save(filename=ROOT_DIR + "\sendMe.xlsx")

    send_mail()                                            


# Change mode between receive and deploy
def change_mode(root):
    if active_mode[0] == 0:
        active_mode[0] = 1
    else:
        active_mode[0] = 0
    for widget in root.winfo_children():
        widget.destroy()

    instantiate_front_end(root)

# take user input and assign equipment to user if valid
def user_submit(t1, t2, root):
    asset = t1.get()
    alias = t2.get()
    if asset.isspace() or alias.isspace() or asset == "" or alias == "":
        n=1
    else:
        update_excel_user(asset, alias)
        root.destroy()

# create user deployment window to deploy equipment to specific user
def user_window(root):
    if active_windows[0] == 0:
        user_win = Toplevel(root)
        user_win.focus()
        user_win.geometry("750x550")
        user_win.title("User Asset Deployment")
        user_win.config(bg="Black", highlightcolor="Orange")

        asset_textbox = tk.Entry(user_win, width=50, borderwidth=1, foreground="orange", background="black", font=("Amazon Ember", 20), justify='center')
        asset_label = tk.Label(user_win, text="Asset Serial Number:", bg="black", fg="white", font=("Amazon Ember", 32), pady=30)
        asset_label.pack()
        asset_textbox.pack()

        user_textbox = tk.Entry(user_win, width=50, borderwidth=1, foreground="orange", background="black", font=("Amazon Ember", 20), justify='center')
        user_label = tk.Label(user_win, text="User Alias:", bg="black", fg="white", font=("Amazon Ember", 32),
                           pady=30)
        user_label.pack()
        user_textbox.pack()

        sub_path = "btn_graphic_submit.png"
        image = Image.open(sub_path)
        sub = ImageTk.PhotoImage(image)
        sub_btn = tk.Button(user_win, text="Submit", width=290, height=140, borderwidth=0, image=sub, compound="center",
                            font=("Amazon Ember", 32), bg="black", fg="white", pady=50)
        sub_btn.config(command=lambda: user_submit(asset_textbox, user_textbox, user_win))
        sub_btn.photo = sub
        sub_btn.pack()


# Send Email function
def send_mail():
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = '[INTERNAL DATA REDACTED]'
    mail.Subject = '[INTERNAL DATA REDACTED]'
    mail.HtmlBody = ''
    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    #(ROOT_DIR + '\sendMe.xlsx')
    mail.Attachments.Add(ROOT_DIR + '\sendMe.xlsx')

    mail.send
    outlook.Application.Quit()

# When user clicks on a location, update form with location data
def location_button_click(alias, label, id):
    active_location_id[0] = id
    temp = len(alias) / 2
    tempx = 130 + 145

    while temp > 0:
        tempx -= 20
        temp -= 1

    label.config(text=alias)
    label.place(x=tempx, y=745)

# Create loading graphic and send update email when submit button is clicked
def submit_button(textbox, root):
    image_path = "loading.png"
    image = Image.open(image_path)
    load = ImageTk.PhotoImage(image)
    load_label = tk.Label(root, image=load, borderwidth=0)
    load_label.photo = load
    load_label.place(x=10, y=550)
    root.update()

    if active_location_id[0] == -1:
        n=0
        load_label.destroy()
        root.update()
    else:
        assets = textbox.get("1.0", 'end-1c')
        assets = blit_inputstr(assets)
        if len(assets) == 0:
            n = 0
            load_label.destroy()
            root.update()
        else:
            textbox.delete('1.0', 'end')
            update_excel(active_location_id[0], assets)

            image_path = "done.png"
            image = Image.open(image_path)
            done = ImageTk.PhotoImage(image)
            load_label.photo = done
            load_label.config(image=done)

# function to split provided string
def blit_inputstr(str):
    assets = str.splitlines()
    return assets

# main function to run program
def main():
    # Instantiate window
    root = tk.Tk()
    root.title("Better DSCS")
    root.attributes("-fullscreen", True)
    root.configure(bg='Black')
    instantiate_front_end(root)
    root.mainloop()

# front end configuration for form
def instantiate_front_end(root):
    active_location_id[0] = -1
    alias_names.clear()

    win_w = root.winfo_screenwidth()
    win_h = root.winfo_screenheight()

    buttons = []

    btn_path = "btn_graphic.PNG"
    image = Image.open(btn_path)
    btn = ImageTk.PhotoImage(image)

    sub_path = "btn_graphic_submit.png"
    image = Image.open(sub_path)
    sub = ImageTk.PhotoImage(image)
    sub_btn = tk.Button(root, text="Submit", width=290, height=140, borderwidth=0, image=sub, compound="center",
                        font=("Amazon Ember", 32), bg="black", fg="white")
    sub_btn.config(command=lambda: submit_button(asset_box, root))
    sub_btn.photo = sub

    logo_2_path = "logopt2.png"
    image = Image.open(logo_2_path)
    logo2 = ImageTk.PhotoImage(image)
    logo2_label = tk.Label(root, image=logo2, borderwidth=0)
    logo2_label.photo = logo2

    if active_mode[0] == 0:
        image_path = "image (6).png"
        image = Image.open(image_path)
        m_ap = ImageTk.PhotoImage(image)
        map_label = tk.Label(root, image=m_ap, borderwidth=0)
        map_label.photo = m_ap

        mode_btn = tk.Button(root, text="Receive Menu", width=290, height=140, borderwidth=0, image=sub,
                             compound="center",
                             font=("Amazon Ember", 32), bg="black", fg="white")
        mode_btn.config(command=lambda: change_mode(root))
        mode_btn.photo = sub
    else:
        image_path = "IT-Hub.png"
        image = Image.open(image_path)
        m_ap = ImageTk.PhotoImage(image)
        map_label = tk.Label(root, image=m_ap, borderwidth=0)
        map_label.photo = m_ap

        mode_btn = tk.Button(root, text="Deploy Menu", width=290, height=140, borderwidth=0, image=sub,
                             compound="center",
                             font=("Amazon Ember", 32), bg="black", fg="white")
        mode_btn.config(command=lambda: change_mode(root))
        mode_btn.photo = sub

    user_btn = tk.Button(root, text="User\nDeployment", width=290, height=140, borderwidth=0, image=sub, compound="center",
                         font=("Amazon Ember", 32), bg="black", fg="white")
    user_btn.config(command=lambda: user_window(root))
    user_btn.photo = sub

    asset_box = tk.Text(root, height=20, width=60, borderwidth=1, foreground="orange", background="black")

    logo_path = "logo.jpg"
    image = Image.open(logo_path)
    logo = ImageTk.PhotoImage(image)
    logo_label = tk.Label(root, image=logo, borderwidth=0)
    logo_label.photo = logo

    location_label = tk.Label(root, text="(empty)")
    location_label.config(font=("Amazon Ember", 30), bg="black", fg="white")

    to_label = tk.Label(root, text="to")
    to_label.config(font=("Amazon Ember", 30), bg="black", fg="white")

    location_index = 0

    if active_mode[0] == 0:
        for location in in_use_locations:

            alias_names.append((apply_alias(location), location_index))
            button = tk.Button(root,image=btn, text=alias_names[location_index][0],
                                 width=145, height=35, fg="white", font=("Amazon Ember", 9), compound="center", borderwidth=0)
            b = location_btn(button, location_index)
            buttons.append(b)
            buttons[location_index].button.image = btn

            location_index += 1
    else:
        for location in spare_locations:

            alias_names.append((apply_alias(location), location_index))
            button = tk.Button(root,image=btn, text=alias_names[location_index][0],
                                 width=145, height=35, fg="white", font=("Amazon Ember", 9), compound="center", borderwidth=0)
            b = location_btn(button, location_index)
            buttons.append(b)
            buttons[location_index].button.image = btn

            location_index += 1

    for b in buttons:
        for alias in alias_names:
            if alias[1] == b.id:
                b.set_config(alias[0], location_label, location_button_click)
                continue

    alias_names.sort()
    xx = win_w - 155
    yy = win_h - 45
    for x, alias in enumerate(alias_names):
        buttons[alias_names[len(alias_names) - 1 - x][1]].button.place(x=xx, y=yy)
        xx -= 155
        if xx < 0:
            xx = win_w - 155
            yy -= 45

    # print(alias_names)
    if win_h >= 1000:
        root.update()
        map_label.place(x=(win_w - 1372), y=(win_h - 733 - (win_h - yy)))
        asset_box.place(x=10, y=(win_h - 733 - (win_h - yy)))
        asset_box.focus()

    tempx = 130
    tempy = yy - 140
    sub_btn.place(x=tempx, y=tempy)

    tempy -= 60
    location_label.place(x=tempx, y=tempy)

    tempx += 100
    tempy -= 100
    to_label.place(x=tempx, y=tempy)

    logo_label.place(x=0, y=0)
    logo2_label.place(x=390, y=0)

    tempx = win_w
    mode_btn.place(x=tempx - 300, y=0)

    user_btn.place(x=tempx - 600, y=0)

# get name of department without auxiliary location data
def apply_alias(name):
    temp_index = len(name) - 1
    while name[temp_index] == '0' or name[temp_index] == '.':
        temp_index -= 1
        continue
    end_index = temp_index + 1
    while name[temp_index] != '.':
        temp_index -= 1
        continue
    start_index = temp_index + 1
    return name[start_index:end_index]

main()
